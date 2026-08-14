"""docs.zip 검수본(파싱 결과_검수본 시트)을 청크 단위 LangChain Document로 만들고,
CLOVA Studio 임베딩으로 Chroma 벡터DB에 적재한다.

문서 준비(prepare_documents)는 API 키 없이도 동작해서 구조 검증이 가능하다.
실제 임베딩 호출(build_vectorstore)은 CLOVASTUDIO_API_KEY가 있어야 한다.
"""

from dataclasses import dataclass
from typing import Optional

import openpyxl
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

HEADER_ROW = 5
DATA_START_ROW = 6

# CLOVA Studio 임베딩(clir-emb-dolphin) 컨텍스트 길이 실측 결과 약 900~950자에서
# "Context length exceeded" 오류가 발생하여, 안전 여유를 두고 이 길이 이상인 청크는
# 여러 조각으로 나눠 임베딩한다 (내용 손실 없이 분할, 절삭하지 않음).
MAX_CHARS_PER_EMBED_CHUNK = 600
EMBED_CHUNK_OVERLAP = 80

# 0-idx 컬럼 위치 (파싱 결과_검수본 시트 기준)
COL_FILE_NUM = 0
COL_FILE_TITLE = 1
COL_FILE_FORMAT = 2
COL_CATEGORY = 3
COL_STORAGE = 4
COL_SECTION = 5
COL_SOURCE_LOC = 6
COL_CHUNK_ID = 7
COL_CHUNK_TEXT = 8
COL_HAS_TABLE = 9
COL_TABLE_TEXT = 10
COL_NOTE = 11
COL_REVIEW_STATUS = 12


@dataclass
class LoadStats:
    documents: int = 0
    skipped_empty: int = 0


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def prepare_documents(xlsm_path: str, sheet_name: str = "파싱 결과_검수본") -> list[Document]:
    """xlsm에서 청크를 읽어 임베딩 대상 Document 리스트로 변환한다. 임베딩 API 호출 없음."""
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws = wb[sheet_name]

    documents: list[Document] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        row = [ws.cell(row=r, column=c + 1).value for c in range(13)]
        chunk_id = _clean(row[COL_CHUNK_ID])
        chunk_text = _clean(row[COL_CHUNK_TEXT])
        if not chunk_id or not chunk_text:
            continue

        table_text = _clean(row[COL_TABLE_TEXT])
        content = chunk_text
        if table_text and table_text not in chunk_text:
            content = f"{chunk_text}\n\n{table_text}"

        metadata = {
            "chunk_id": chunk_id,
            "file_number": row[COL_FILE_NUM],
            "file_title": _clean(row[COL_FILE_TITLE]),
            "file_format": _clean(row[COL_FILE_FORMAT]),
            "category": _clean(row[COL_CATEGORY]),
            "section": _clean(row[COL_SECTION]),
            "source_location": _clean(row[COL_SOURCE_LOC]),
            "has_table": _clean(row[COL_HAS_TABLE]),
        }
        documents.append(Document(page_content=content, metadata=metadata))

    return _split_long_documents(documents)


def _split_long_documents(documents: list[Document]) -> list[Document]:
    """MAX_CHARS_PER_EMBED_CHUNK를 넘는 문서를 여러 조각으로 나눈다 (내용 절삭 없음)."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=MAX_CHARS_PER_EMBED_CHUNK,
        chunk_overlap=EMBED_CHUNK_OVERLAP,
        separators=["\n\n", "\n", " ", ""],
    )

    result: list[Document] = []
    for doc in documents:
        if len(doc.page_content) <= MAX_CHARS_PER_EMBED_CHUNK:
            result.append(doc)
            continue

        parts = splitter.split_text(doc.page_content)
        base_chunk_id = doc.metadata["chunk_id"]
        for i, part in enumerate(parts, start=1):
            metadata = dict(doc.metadata)
            metadata["chunk_id"] = f"{base_chunk_id}_part{i}"
            metadata["parent_chunk_id"] = base_chunk_id
            metadata["part_index"] = i
            metadata["part_count"] = len(parts)
            result.append(Document(page_content=part, metadata=metadata))

    return result


class _RateLimitedEmbeddings:
    """테스트 키의 요청 제한(429) 대응 — 호출 사이 지연 + 429 발생시 지수 백오프 재시도."""

    def __init__(self, inner, delay_seconds: float = 1.5, max_retries: int = 6):
        self._inner = inner
        self._delay = delay_seconds
        self._max_retries = max_retries

    def _call_with_retry(self, text: str) -> list[float]:
        import time

        from openai import RateLimitError

        wait = self._delay
        for attempt in range(self._max_retries):
            try:
                result = self._inner.embed_query(text)
                time.sleep(self._delay)
                return result
            except RateLimitError:
                if attempt == self._max_retries - 1:
                    raise
                time.sleep(wait)
                wait *= 2
        raise RuntimeError("unreachable")

    def embed_documents(self, texts: list[str]) -> list[list[float]]:
        return [self._call_with_retry(t) for t in texts]

    def embed_query(self, text: str) -> list[float]:
        return self._call_with_retry(text)


def build_vectorstore(
    xlsm_path: str,
    persist_dir: str,
    collection_name: str = "pension_docs",
    embedding_model: str = "clir-emb-dolphin",
    request_delay_seconds: float = 1.5,
    batch_size: int = 20,
):
    """Document를 준비하고 CLOVA Studio로 임베딩하여 Chroma에 적재한다. API 키 필요.

    테스트 키의 낮은 요청 제한(rate limit)을 감안해 호출 사이에 지연을 두고,
    작은 배치 단위로 나눠 진행 상황을 print로 보여준다.
    """
    from langchain_chroma import Chroma

    from src.agents.llm import get_embeddings

    documents = prepare_documents(xlsm_path)
    raw_embeddings = get_embeddings(embedding_model)
    embeddings = _RateLimitedEmbeddings(raw_embeddings, delay_seconds=request_delay_seconds)

    vectorstore = Chroma(
        collection_name=collection_name,
        embedding_function=embeddings,
        persist_directory=persist_dir,
    )

    # 중간에 실패해도 재실행 시 이미 적재된 chunk_id는 건너뛰어 이어서 진행한다.
    existing_ids = set(vectorstore.get(include=[])["ids"])
    documents = [d for d in documents if d.metadata["chunk_id"] not in existing_ids]
    if existing_ids:
        print(f"  이미 적재된 {len(existing_ids)}건 건너뜀, 남은 {len(documents)}건 진행")

    total = len(documents)
    for start in range(0, total, batch_size):
        batch = documents[start:start + batch_size]
        ids = [doc.metadata["chunk_id"] for doc in batch]
        vectorstore.add_documents(documents=batch, ids=ids)
        print(f"  적재 진행: {min(start + batch_size, total)}/{total}")

    final_count = vectorstore._collection.count()
    return vectorstore, LoadStats(documents=final_count)
