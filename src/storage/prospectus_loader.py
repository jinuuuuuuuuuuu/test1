"""검수 완료된 투자설명서 xlsm(검수본 시트)을 fund_master/fund_class SQLite DB로 적재한다."""

import datetime
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl

from src.storage.schema import connect

HEADER_ROW = 5
DATA_START_ROW = 6

# 0-idx 컬럼 위치 (검수본 시트 기준)
COL = {
    "source_file": 0, "fund_name": 1, "product_code": 2, "manager_name": 3,
    "base_date": 4, "prospectus_effective_date": 5, "risk_grade": 6, "issue_type": 7,
    "investment_objective": 8, "investment_strategy": 9, "fund_category": 10,
    "class_name": 11, "sales_channel": 12, "total_expense_ratio": 13,
    "cost_3y_per_10m_krw": 14, "return_asof_date": 15, "return_1y": 16,
    "return_3y": 17, "return_since_inception": 18, "volatility_1y": 19,
    "volatility_3y": 20, "volatility_since_inception": 21, "benchmark": 22,
    "manager_person_names": 23, "manager_birth_years": 24, "manager_titles": 25,
    "manager_years_experience": 26, "peer_avg_return_1y": 27, "purchase_rule": 28,
    "redemption_rule": 29, "redemption_payment_rule": 30, "redemption_fee": 31,
    "tax_note": 32, "conversion_available": 33, "inception_date": 34,
    "extraction_source": 35, "extraction_note": 36, "review_status": 37,
}

MASTER_FIELDS = [
    "source_file", "fund_name", "manager_name", "base_date", "prospectus_effective_date",
    "risk_grade", "issue_type", "investment_objective", "investment_strategy", "fund_category",
    "benchmark", "manager_person_names", "manager_birth_years", "manager_titles",
    "manager_years_experience", "peer_avg_return_1y", "purchase_rule", "redemption_rule",
    "redemption_payment_rule",
]

CLASS_FIELDS = [
    "class_name", "sales_channel", "total_expense_ratio", "cost_3y_per_10m_krw",
    "return_asof_date", "return_1y", "return_3y", "return_since_inception",
    "volatility_1y", "volatility_3y", "volatility_since_inception", "redemption_fee",
    "tax_note", "conversion_available", "inception_date", "extraction_source",
    "extraction_note", "review_status",
]

NUMERIC_FIELDS = {
    "total_expense_ratio", "cost_3y_per_10m_krw", "return_1y", "return_3y",
    "return_since_inception", "volatility_1y", "volatility_3y", "volatility_since_inception",
}


def _clean(value: Any, numeric: bool = False) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        if text.upper() == "NULL" or text == "":
            return None
        if numeric:
            try:
                return float(text)
            except ValueError:
                return None
        return text
    if numeric and isinstance(value, (int, float)):
        return float(value)
    return value


@dataclass
class LoadStats:
    funds: int = 0
    classes: int = 0
    skipped_rows: int = 0


def load_prospectus_xlsm(xlsm_path: str, db_path: str, sheet_name: str = "투자설명서_파싱_검수본") -> LoadStats:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws = wb[sheet_name]

    conn = connect(db_path)
    stats = LoadStats()
    seen_master: set[str] = set()

    with conn:
        conn.execute("DELETE FROM fund_class")
        conn.execute("DELETE FROM fund_master")

        for r in range(DATA_START_ROW, ws.max_row + 1):
            row = [ws.cell(row=r, column=c + 1).value for c in range(len(COL))]
            product_code = _clean(row[COL["product_code"]])
            if not product_code:
                stats.skipped_rows += 1
                continue

            if product_code not in seen_master:
                master_vals = {f: _clean(row[COL[f]]) for f in MASTER_FIELDS}
                conn.execute(
                    f"INSERT INTO fund_master (product_code, {', '.join(MASTER_FIELDS)}) "
                    f"VALUES (?, {', '.join(['?'] * len(MASTER_FIELDS))})",
                    [product_code] + [master_vals[f] for f in MASTER_FIELDS],
                )
                seen_master.add(product_code)
                stats.funds += 1

            class_vals = {f: _clean(row[COL[f]], numeric=(f in NUMERIC_FIELDS)) for f in CLASS_FIELDS}
            conn.execute(
                f"INSERT INTO fund_class (product_code, {', '.join(CLASS_FIELDS)}) "
                f"VALUES (?, {', '.join(['?'] * len(CLASS_FIELDS))})",
                [product_code] + [class_vals[f] for f in CLASS_FIELDS],
            )
            stats.classes += 1

    conn.close()
    return stats
