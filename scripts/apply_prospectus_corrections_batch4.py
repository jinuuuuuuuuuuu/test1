"""투자설명서 18~23번 파일 검증 결과 반영."""

import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서.xlsm"

wb = openpyxl.load_workbook(PATH, keep_vba=True)
ws = wb["투자설명서_파싱"]


def find_rows(code):
    return [row for row in ws.iter_rows(min_row=6) if row[2].value == code]


def find_row_by_class(code, cls):
    for row in ws.iter_rows(min_row=6):
        if row[2].value == code and row[11].value == cls:
            return row
    raise ValueError(f"{code}/{cls} not found")


def fill_common(row, 매입, 환매, 지급, 수수료, 과세, note_suffix=""):
    if row[28].value in (None, "NULL", ""):
        row[28].value = 매입
    if row[29].value in (None, "NULL", ""):
        row[29].value = 환매
    if row[30].value in (None, "NULL", ""):
        row[30].value = 지급
    if 수수료 is not None and row[31].value in (None, "NULL", ""):
        row[31].value = 수수료
    if row[32].value in (None, "NULL", ""):
        row[32].value = 과세
    if note_suffix:
        row[36].value = ((row[36].value or "") + " " + note_suffix).strip()
    row[37].value = "검수완료"


# ── 18. KR5113420069 ─────────────────────────────────────────────
for r in find_rows("KR5113420069"):
    fill_common(
        r,
        "17시 이전: 납입영업일로부터 2영업일(D+1) 공고 기준가격 적용 매입 / 17시 경과 후: 3영업일(D+2) 기준가격 적용 매입",
        "오후5시 이전: 환매청구일로부터 3영업일(D+2) 기준가격 적용, 3영업일에 세금 등 공제 후 지급 / 오후5시 경과 후: 4영업일(D+3) 기준가격 적용, 4영업일에 세금 등 공제 후 지급",
        "위 환매방법 참고",
        None,  # 이미 원문 확인상 항목별로 다르므로 강제로 채우지 않음(아래서 직접 처리)
        "수익자: 15.4%(지방소득세포함) 원천징수, 기준금액 초과시 종합과세. 국내상장주식 매매·평가손익 과세제외 유의. 퇴직연금계좌: 원천징수 없이 수령시 관련세법 과세(일반 투자신탁과 세율 상이)",
        "원문 대조하여 NULL 필드 보완.",
    )
row_ce = find_row_by_class("KR5113420069", "C-e")
row_ce[16].value = 3.41
row_ce[17].value = 6.31
row_ce[18].value = 3.55
row_ce[36].value = (row_ce[36].value or "") + (
    " ⚠️수익률 NULL→실측치로 보완: 원문 '가.연평균수익률'표 수수료미징구-온라인(C-e) 행에서 1년=3.41, 3년=6.31, 설정이후=3.55 확인."
)
row_ce[37].value = "검수완료(수정)"

# ── 19. KR5113450111 ─────────────────────────────────────────────
for r in find_rows("KR5113450111"):
    fill_common(
        r,
        "오후3시30분 이전: 납입영업일로부터 2영업일(D+1) 공고 기준가격 적용 매입 / 오후3시30분 경과 후: 3영업일(D+2) 기준가격 적용 매입",
        "오후3시30분 이전: 환매청구일로부터 2영업일(D+1) 기준가격 적용, 4영업일(D+3)에 세금 등 공제 후 지급 / 오후3시30분 경과 후: 3영업일(D+2) 기준가격 적용, 4영업일(D+3)에 세금 등 공제 후 지급",
        "위 환매방법 참고",
        "없음",
        "연금저축계좌 가입자: 환매시점 별도과세 없음, 인출시 관련세법에 따라 과세(일반 투자신탁과 세율 상이)",
        "원문 대조하여 NULL 필드 보완. 변동성 NULL 처리는 원문 확인 결과 타당함(원문 PDF 자체의 '수익률변동성' 행이 모펀드 수익률 행과 동일 숫자로 중복 기재된 구조적 결함이 있어 신뢰 불가 — 팀의 기존 판단이 맞음).",
    )
row_c = find_row_by_class("KR5113450111", "C")
row_c[16].value = -3.19
row_c[17].value = 0.23
row_c[18].value = 7.22
row_c[37].value = "검수완료"  # 값 자체는 기존과 동일(재확인)

row_ce = find_row_by_class("KR5113450111", "C-e")
row_ce[16].value = -2.73
row_ce[17].value = 0.71
row_ce[18].value = 3.40
row_ce[36].value = (row_ce[36].value or "") + (
    " ⚠️수익률 NULL→실측치로 보완: 원문 '가.연평균수익률'표 (C-e) 행에서 1년=-2.73, 3년=0.71, 설정이후=3.40 확인."
)
row_ce[37].value = "검수완료(수정)"

# ── 20. KR5113450401 ─────────────────────────────────────────────
for r in find_rows("KR5113450401"):
    fill_common(
        r,
        "오후3시30분 이전: 납입영업일로부터 2영업일(D+1) 공고 기준가격 적용 매입 / 오후3시30분 경과 후: 3영업일(D+2) 기준가격 적용 매입",
        "오후3시30분 이전: 환매청구일로부터 2영업일(D+1) 기준가격 적용 / 오후3시30분 경과 후: 3영업일(D+2) 기준가격 적용, 4영업일(D+3)에 세금 등 공제 후 지급",
        "4영업일(D+3)에 세금 등 공제 후 지급",
        "없음",
        "수익자: 15.4%(지방소득세포함) 원천징수, 기준금액 초과시 종합과세. 국내상장주식 매매·평가손익 과세제외 유의",
        "원문 대조하여 NULL 필드 보완.",
    )
row_c = find_row_by_class("KR5113450401", "C")
row_c[19].value = None  # 변동성1년 -> 아래서 재설정
row_c[19].value = "NULL"
row_c[20].value = "NULL"
row_c[21].value = "NULL"
row_c[36].value = (row_c[36].value or "") + (
    " ⚠️변동성 값 재검토: 기존 29.31/16.15/8.36은 원문 '수익률변동성(%)' 행 값이나, 이 값은 원문 PDF 표 구조상 "
    "'종류모(펀드 전체)'의 수익률 행과 완전히 동일한 숫자가 그대로 중복 기재된 것으로 확인됨(진짜 변동성 수치가 아님, "
    "KR5113450111에서 팀이 이미 동일 패턴을 발견해 NULL 처리한 것과 같은 원문 자체의 결함). "
    "클래스별 실제 변동성 수치가 별도로 존재하지 않아 NULL로 되돌림."
)
row_c[37].value = "검수완료(수정: 변동성 NULL로 정정)"

row_ce = find_row_by_class("KR5113450401", "C-e")
row_ce[16].value = 28.07
row_ce[17].value = 15.01
row_ce[18].value = 7.31
row_ce[19].value = "NULL"
row_ce[20].value = "NULL"
row_ce[21].value = "NULL"
row_ce[36].value = (row_ce[36].value or "") + (
    " ⚠️수익률 NULL→실측치로 보완: 1년=28.07, 3년=15.01, 설정이후=7.31 (원문 '가.연평균수익률'표 (C-e) 행). "
    "⚠️변동성 값 재검토: 위 C클래스와 동일한 이유로 NULL로 되돌림(원문 표 구조 결함)."
)
row_ce[37].value = "검수완료(수정)"

# ── 21/22. KR5113470030 / KR5113470031 (동일 PDF 중복) ──────────
for code in ("KR5113470030", "KR5113470031"):
    for r in find_rows(code):
        fill_common(
            r,
            "오후3시30분 이전: 납입영업일로부터 2영업일(D+1) 공고 기준가격 적용 매입 / 오후3시30분 경과 후: 3영업일(D+2) 기준가격 적용 매입",
            "오후3시30분 이전: 환매청구일로부터 2영업일(D+1) 기준가격 적용 / 오후3시30분 경과 후: 3영업일(D+2) 기준가격 적용, 4영업일(D+3)에 세금 등 공제 후 지급",
            "4영업일(D+3)에 세금 등 공제 후 지급",
            None,
            "수익자: 15.4%(지방소득세포함) 원천징수, 기준금액 초과시 종합과세. 국내상장주식 매매·평가손익 과세제외 유의",
            "⚠️중복 의심: R2_KR5113470030.pdf와 R2_KR5113470031.pdf가 MD5 해시까지 동일한 완전 동일 파일. "
            "두 상품코드가 실제로 서로 다른 등록증권인지 원본 데이터 중복 오류인지 이 문서만으로 판단 불가 — "
            "원본 데이터셋 제공처(zip) 재확인 필요. (동일 패턴이 KR5113420013/KR5113420015에서도 발견됨 — 원본 zip 자체의 "
            "체계적 문제일 가능성 있음)",
        )
    row_c = find_row_by_class(code, "C")
    row_c[16].value = 74.00
    row_c[17].value = 26.53
    row_c[18].value = 7.20
    row_c[36].value = (row_c[36].value or "") + " (수익률 값 자체는 기존 파싱값과 일치, 원문 대조로 재확인함.)"
    row_c[37].value = "검수완료(중복 플래그)"

    row_ce = find_row_by_class(code, "C-e")
    row_ce[16].value = 74.80
    row_ce[17].value = 27.15
    row_ce[18].value = 7.68
    row_ce[36].value = (row_ce[36].value or "") + (
        " ⚠️수익률 NULL→실측치로 보완: 1년=74.80, 3년=27.15, 설정이후=7.68 (원문 '가.연평균수익률'표 (C-e) 행)."
    )
    row_ce[37].value = "검수완료(수정, 중복 플래그)"

# ── 23. KR5114420016 ─────────────────────────────────────────────
for r in find_rows("KR5114420016"):
    fill_common(
        r,
        "오후5시 이전: 납입영업일로부터 2영업일(D+1) 공고 기준가격 적용 매입 / 오후5시 경과 후: 3영업일(D+2) 기준가격 적용 매입",
        "오후5시 이전: 환매청구일로부터 3영업일(D+2) 기준가격 적용, 3영업일에 세금 등 공제 후 지급 / 오후5시 경과 후: 4영업일(D+3) 기준가격 적용, 4영업일에 세금 등 공제 후 지급",
        "위 환매방법 참고",
        None,
        "연금저축계좌 가입자: 환매시점 별도과세 없음, 인출시 관련세법에 따라 과세(일반 투자신탁과 세율 상이)",
        "원문 대조하여 NULL 필드 보완.",
    )
row_cp = find_row_by_class("KR5114420016", "C-P")
row_cp[16].value = 8.93
row_cp[17].value = 4.79
row_cp[18].value = 0.94
row_cp[34].value = "2016-07-14"
row_cp[36].value = (row_cp[36].value or "") + (
    " ⚠️최초설정일 정정: 2005-04-26(펀드 전체 설정일) → 2016-07-14(원문 '클래스별 설정일' 표 및 (9)번 실적 상세표 확인). "
    "⚠️수익률 NULL→실측치로 보완: 1년=8.93, 3년=4.79, 설정이후=0.94."
)
row_cp[37].value = "검수완료(수정)"

row_cpe = find_row_by_class("KR5114420016", "C-Pe")
row_cpe[16].value = 9.13
row_cpe[17].value = 4.98
row_cpe[18].value = 1.76
row_cpe[34].value = "2017-08-02"
row_cpe[36].value = (row_cpe[36].value or "") + (
    " ⚠️최초설정일 정정: 2005-04-26(펀드 전체 설정일) → 2017-08-02(원문 '클래스별 설정일' 표 및 (10)번 실적 상세표 확인). "
    "⚠️수익률 NULL→실측치로 보완: 1년=9.13, 3년=4.98, 설정이후=1.76."
)
row_cpe[37].value = "검수완료(수정)"

wb.save(PATH)
print("Batch4 (18-23) saved.")
