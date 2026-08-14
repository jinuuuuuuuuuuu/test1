import io
import openpyxl

TRACKER_PATH = r"C:\Users\kevin\Downloads\docs 수정.xlsm"
SRC_PATH = r"C:\Users\kevin\pension-agent\data\raw\docs\docs_renamed\doc34.xlsx"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_doc34_diff.txt"

wb_src = openpyxl.load_workbook(SRC_PATH, data_only=True)
ws_src = wb_src["Sheet1"]

src_rows = []
for row in ws_src.iter_rows(min_row=7, min_col=2, max_col=3, values_only=True):
    code, reason = row
    if code is None:
        continue
    src_rows.append((str(code).strip(), str(reason).strip() if reason else ""))

wb_t = openpyxl.load_workbook(TRACKER_PATH, data_only=True)
ws_t = wb_t["파싱 결과"]

diffs = []
count = 0
for row in ws_t.iter_rows(min_row=6, values_only=True):
    if row[0] is None:
        continue
    try:
        num = int(float(row[0]))
    except (TypeError, ValueError):
        continue
    if num != 34:
        continue
    count += 1
    chunk_id = row[7]
    chunk_text = row[8] or ""
    try:
        n = int(chunk_id.split("chunk")[1])
    except Exception:
        diffs.append(f"{chunk_id}: cannot parse chunk number")
        continue
    if n - 1 >= len(src_rows):
        diffs.append(f"{chunk_id}: no matching src row index {n}")
        continue
    code, reason = src_rows[n - 1]
    problems = []
    # normalize whitespace for comparison
    norm_chunk = " ".join(chunk_text.split())
    norm_code = " ".join(code.split())
    norm_reason = " ".join(reason.split())
    if norm_code not in norm_chunk:
        problems.append(f"[코드] SRC='{code}' NOT FOUND in chunk")
    if norm_reason and norm_reason not in norm_chunk:
        problems.append(f"[사유] SRC='{reason[:80]}...' NOT FOUND verbatim in chunk")
    if problems:
        diffs.append(f"{chunk_id} (src row {n}, code={code}):\n" + "\n".join(problems))

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"총 tracker doc34 청크 수: {count}, 원본 코드 행 수: {len(src_rows)}\n\n")
    if diffs:
        f.write(f"불일치 {len(diffs)}건:\n\n")
        f.write("\n\n".join(diffs))
    else:
        f.write("불일치 없음 - 전체 일치\n")
