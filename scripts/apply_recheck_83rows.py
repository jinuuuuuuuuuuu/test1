# -*- coding: utf-8 -*-
import io
import re
import datetime
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
LOG_OUT = r"C:\Users\kevin\pension-agent\docs_check\_apply_recheck_83rows_log.txt"

COL_INDEX = {
    "파일명": 0, "집합투자기구명칭": 1, "상품코드": 2, "집합투자업자명칭": 3, "작성기준일": 4,
    "증권신고서효력발생일": 5, "투자위험등급": 6, "모집(매출)증권 종류/총액": 7, "투자목적": 8, "투자전략": 9,
    "상품분류": 10, "판매클래스": 11, "판매방식": 12, "총보수·비용(%)": 13, "1000만원투자시_3년총비용(천원)": 14,
    "수익률기준일": 15, "수익률_최근1년(%)": 16, "수익률_최근3년(%)": 17, "수익률_설정일이후(%)": 18,
    "변동성_최근1년(%)": 19, "변동성_최근3년(%)": 20, "변동성_설정일이후(%)": 21, "비교지수(BM)": 22,
    "책임운용역_성명": 23, "책임운용역_생년": 24, "책임운용역_직위": 25, "책임운용역_운용경력년수": 26,
    "동종집합투자기구_운용사최근1년수익률(%)": 27, "매입기준": 28, "환매기준": 29, "환매대금지급": 30,
    "환매수수료": 31, "과세특징": 32, "전환가능여부(전환형)": 33, "최초설정일": 34, "근거위치": 35,
    "검수메모": 36, "검수상태": 37,
}

NUMERIC_COLS = {
    "총보수·비용(%)", "1000만원투자시_3년총비용(천원)", "수익률_최근1년(%)", "수익률_최근3년(%)",
    "수익률_설정일이후(%)", "변동성_최근1년(%)", "변동성_최근3년(%)", "변동성_설정일이후(%)",
    "동종집합투자기구_운용사최근1년수익률(%)",
}

DATE_COLS = {"최초설정일"}


def parse_labeled_values(text, labels):
    if not text:
        return {}
    positions = []
    for label in labels:
        idx = text.find(label + ":")
        if idx == -1:
            continue
        positions.append((idx, label))
    positions.sort()
    result = {}
    for i, (idx, label) in enumerate(positions):
        start = idx + len(label) + 1
        end = positions[i + 1][0] if i + 1 < len(positions) else len(text)
        val = text[start:end].strip()
        result[label] = val
    return result


def coerce_value(label, raw):
    raw = (raw or "").strip()
    if raw.upper() == "NULL":
        return "NULL"
    if label in NUMERIC_COLS:
        try:
            return float(raw)
        except ValueError:
            return raw
    if label in DATE_COLS:
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", raw)
        if m:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return raw
    return raw


wb = openpyxl.load_workbook(PATH, keep_vba=True)
recheck_ws = wb["투자설명서_재검수_추가수정필요"]
target_ws = wb["투자설명서_파싱_검수본"]

# 상품코드+클래스 -> 실제 행 매핑 (검수본 현재 상태 기준)
code_class_to_row = {}
for r in range(6, target_ws.max_row + 1):
    code = target_ws.cell(row=r, column=3).value
    cls = target_ws.cell(row=r, column=12).value
    if code:
        code_class_to_row[(code, cls)] = r

applied = []
skipped = []
delete_instructions = []

for r in range(5, recheck_ws.max_row + 1):
    vals = [recheck_ws.cell(row=r, column=c).value for c in range(1, 13)]
    if vals[0] is None:
        continue
    orig_row, fname, code, cls, err_cols_raw, cur_raw, rec_raw, err_type, direction, source, method, confidence = vals

    err_cols = [c.strip() for c in (err_cols_raw or "").split("\n") if c.strip()]

    if err_cols == ["행처리"]:
        delete_instructions.append((code, cls, direction, rec_raw))
        continue

    rec_vals = parse_labeled_values(rec_raw, err_cols)

    target_row = code_class_to_row.get((code, cls))
    if target_row is None:
        skipped.append((code, cls, "매칭되는 검수본 행 없음"))
        continue

    row_cells = target_ws[target_row]
    note_parts = []
    for label in err_cols:
        if label not in COL_INDEX:
            skipped.append((code, cls, f"알 수 없는 컬럼명: {label}"))
            continue
        if label not in rec_vals:
            skipped.append((code, cls, f"권고값 파싱 실패: {label}"))
            continue
        col_idx = COL_INDEX[label]
        new_val = coerce_value(label, rec_vals[label])
        row_cells[col_idx].value = new_val
        note_parts.append(label)

    if note_parts:
        note = f"⚠️팀원 재검수(투자설명서 (2).xlsm 재검수_추가수정필요 row{r})로 정정: {', '.join(note_parts)}. 근거: {source}"
        existing_note = row_cells[36].value or ""
        row_cells[36].value = (existing_note + " " + note).strip()
        row_cells[37].value = "검수완료(재검수 반영)"
        applied.append((code, cls, note_parts))

# 중복행 삭제 처리 (최후에, 상품코드+클래스로 재확인 후 삭제)
deleted = []
for code, cls, direction, rec_raw in delete_instructions:
    target_row = code_class_to_row.get((code, cls))
    if target_row is None:
        skipped.append((code, cls, "삭제대상 행을 찾지 못함"))
        continue
    target_ws.delete_rows(target_row, 1)
    deleted.append((code, cls, target_row, direction))

wb.save(PATH)

with io.open(LOG_OUT, "w", encoding="utf-8") as f:
    f.write(f"반영됨: {len(applied)}건\n")
    for code, cls, cols in applied:
        f.write(f"  {code} ({cls}): {cols}\n")
    f.write(f"\n삭제됨: {len(deleted)}건\n")
    for code, cls, row, direction in deleted:
        f.write(f"  {code} ({cls}) row{row}: {direction}\n")
    f.write(f"\n스킵됨: {len(skipped)}건\n")
    for code, cls, reason in skipped:
        f.write(f"  {code} ({cls}): {reason}\n")

print("done. applied:", len(applied), "deleted:", len(deleted), "skipped:", len(skipped))
