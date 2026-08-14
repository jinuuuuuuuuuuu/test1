import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_current_17rows.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

TARGETS = [
    ("KR5147430065", "C"), ("KR5147430065", "Ce"),
    ("KR5153420022", "C"), ("KR5153420022", "C-e"),
    ("KR5153420063", "C-P"), ("KR5153420063", "C-P2"),
    ("KR5153420318", "C"),
    ("KR5153450209", "C"), ("KR5153450209", "C-P2e"),
    ("KR5153450250", "C"), ("KR5153450250", "C-P2e"),
    ("KR5153450268", "C1"), ("KR5153450268", "C-e"),
    ("KR5153450431", "C"), ("KR5153450431", "C-e"),
    ("KR5153450658", "C1"), ("KR5153450658", "C-e"),
]

with io.open(OUT, "w", encoding="utf-8") as f:
    for r in range(6, ws.max_row + 1):
        code = ws.cell(row=r, column=3).value
        cls = ws.cell(row=r, column=12).value
        if (code, cls) in TARGETS:
            f.write(f"row{r}: {code} ({cls}) 1y={ws.cell(row=r,column=17).value} 3y={ws.cell(row=r,column=18).value} "
                    f"since={ws.cell(row=r,column=19).value} 최초설정일={ws.cell(row=r,column=35).value}\n")
