import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_prospectus2_inspect.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"시트 목록: {wb.sheetnames}\n\n")
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        f.write(f"===== {sheetname} (dims={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) =====\n")
