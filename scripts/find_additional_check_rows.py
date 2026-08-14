import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_additional_check_rows.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

rows = []
for r in range(6, ws.max_row + 1):
    status = ws.cell(row=r, column=38).value
    if status and "추가확인" in str(status):
        code = ws.cell(row=r, column=3).value
        cls = ws.cell(row=r, column=12).value
        note = ws.cell(row=r, column=37).value
        rows.append((r, code, cls, status, note))

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"'추가확인' 포함 상태인 행: {len(rows)}건\n\n")
    for r, code, cls, status, note in rows:
        f.write(f"row{r}: {code} ({cls}) 검수상태={status}\n  메모: {note}\n\n")
