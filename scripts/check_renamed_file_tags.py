import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]renamed폴더.xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_renamed_file_check.txt"

TARGETS = ["doc15_chunk04", "doc41_chunk03", "doc41_chunk04"]

wb = openpyxl.load_workbook(PATH, data_only=True)

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"시트 목록: {wb.sheetnames}\n\n")
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        f.write(f"===== 시트: {sheetname} =====\n")
        found_any = False
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                if cell.value in TARGETS:
                    found_any = True
                    r = cell.row
                    rowvals = [ws.cell(row=r, column=c).value for c in range(1, 14)]
                    f.write(f"row {r}: {rowvals}\n")
        if not found_any:
            f.write("(대상 chunk_id 못 찾음)\n")
        f.write("\n")
