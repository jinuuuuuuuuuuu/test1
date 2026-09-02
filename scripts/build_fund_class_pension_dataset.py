"""Build the canonical Cost Guard pension-class dataset.

This script does not touch prospectus.db.  It turns the improved parser output
into reviewable CSV artifacts, compares it with the manually reviewed workbook,
and reports how much lower-cost-class coverage the clean parser output gives.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import tempfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pymupdf

try:
    from scripts.parse_fund_class_pension import collect_clean_rows_from_text_dir
except ModuleNotFoundError:  # pragma: no cover - direct script execution path
    from parse_fund_class_pension import collect_clean_rows_from_text_dir

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = REPO_ROOT / "data" / "processed"
DEFAULT_XLSX = Path("/Users/dapanman/Downloads/Cost_Guard_fund_class_v2_검증보정.xlsx")
DEFAULT_PDF_ROOT = Path("/Users/dapanman/Desktop/2.연금/투자설명서")

ALLOWED_ACCOUNT_TYPES = {"연금저축", "퇴직연금/IRP"}
ALLOWED_CHANNELS = {"오프라인", "온라인", "온라인슈퍼", "직판"}
STANDARD_CHANNELS = {"오프라인", "온라인"}
GOLDEN_VIP = {
    "C-P": 1.66,
    "C-PE": 1.26,
    "C-P2": 1.56,
    "C-P2E": 1.21,
    "S-P": 1.13,
    "S-P2": 1.12,
}
COST_METRIC_FIELDS = (
    "synthetic_total_expense_ratio",
    "total_expense_ratio",
)
PRIMARY_REVIEW_FIELDS = {
    "total_expense_ratio",
    "synthetic_total_expense_ratio",
    "cost_3y_per_10m_krw",
}
REVIEW_STATUSES = {
    "PARSER_CORRECT",
    "REFERENCE_CORRECT",
    "BOTH_WRONG",
    "AMBIGUOUS",
    "SOURCE_CONFLICT",
}
REVIEW_FIELDS = PRIMARY_REVIEW_FIELDS | {"channel", "account_type"}
REVIEW_FIELDNAMES = [
    "product_code",
    "class_code",
    "field",
    "parser_value",
    "reference_value",
    "review_status",
    "reviewed_value",
    "source_page",
    "note",
]


def normalize_class_code(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", "", text).split("(")[0].upper()


def product_code_from_filename(value: Any) -> str | None:
    if value is None:
        return None
    match = re.search(r"KR[0-9A-Z]+", str(value).upper())
    return match.group(0) if match else None


def _float_or_none(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "").replace(",", "")
    if not text or text in {"미확인", "-", "없음"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_pdf_files(pdf_root: Path) -> list[tuple[str, Path]]:
    results: list[tuple[str, Path]] = []
    for code_dir in sorted(p for p in pdf_root.iterdir() if p.is_dir()):
        pdfs = sorted(p for p in code_dir.iterdir() if p.suffix.lower() == ".pdf")
        if pdfs:
            results.append((code_dir.name, pdfs[0]))
    return results


def extract_pdf_text(pdf_path: Path) -> str:
    parts: list[str] = []
    with pymupdf.open(str(pdf_path)) as doc:
        for index, page in enumerate(doc, start=1):
            parts.append(f"\n===== PAGE {index}/{len(doc)} =====\n")
            parts.append(page.get_text() or "")
    return "".join(parts)


def build_text_cache(pdf_root: Path, text_dir: Path) -> None:
    text_dir.mkdir(parents=True, exist_ok=True)
    for code, pdf_path in find_pdf_files(pdf_root):
        out = text_dir / f"{code}.txt"
        if out.exists():
            continue
        out.write_text(extract_pdf_text(pdf_path), encoding="utf-8")


def parser_rows_from_pdf_root(pdf_root: Path) -> tuple[list[dict], list[dict]]:
    with tempfile.TemporaryDirectory(prefix="fund-class-pension-") as tmp:
        text_dir = Path(tmp)
        build_text_cache(pdf_root, text_dir)
        return collect_clean_rows_from_text_dir(str(text_dir))


def read_reference_rows(
    xlsx_path: Path,
    sheet_index: int = 1,
    *,
    cost_guard_usable_only: bool = True,
) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[sheet_index]]
    headers = [cell.value for cell in ws[1]]
    idx = {name: i for i, name in enumerate(headers)}
    required = [
        "투자설명서_파일명",
        "판매클래스_원문",
        "판매클래스_정규화키",
        "계좌유형_v2",
        "판매채널_v2",
        "총보수비용_원본(%)",
        "3년총비용_v2(천원)",
        "검증상태",
        "CostGuard_사용가능",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(f"검증보정 엑셀 필수 컬럼 없음: {', '.join(missing)}")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        product_code = product_code_from_filename(row[idx["투자설명서_파일명"]])
        account_type = row[idx["계좌유형_v2"]]
        usable = row[idx["CostGuard_사용가능"]]
        if not product_code:
            continue
        if cost_guard_usable_only and (account_type not in ALLOWED_ACCOUNT_TYPES or usable != "Y"):
            continue
        class_code = normalize_class_code(row[idx["판매클래스_정규화키"]] or row[idx["판매클래스_원문"]])
        if not class_code:
            continue
        rows.append({
            "product_code": product_code,
            "source_file": row[idx["투자설명서_파일명"]],
            "class_code": class_code,
            "account_type": account_type,
            "channel": row[idx["판매채널_v2"]],
            "total_expense_ratio": _float_or_none(row[idx["총보수비용_원본(%)"]]),
            "synthetic_total_expense_ratio": None,
            "cost_3y_per_10m_krw": _float_or_none(row[idx["3년총비용_v2(천원)"]]),
            "total_expense_source_page": "",
            "synthetic_expense_source_page": "",
            "cost_3y_source_page": "",
            "class_label": row[idx["판매클래스_원문"]],
            "reference_status": row[idx["검증상태"]],
            "cost_guard_usable": usable,
        })
    return rows


def normalize_parser_rows(rows: list[dict]) -> list[dict]:
    normalized = []
    for row in rows:
        normalized.append({
            **row,
            "class_code": normalize_class_code(row["class_code"]),
            "source_file": str(row["source_file"]),
            "total_expense_ratio": _float_or_none(row.get("total_expense_ratio")),
            "synthetic_total_expense_ratio": _float_or_none(row.get("synthetic_total_expense_ratio")),
            "cost_3y_per_10m_krw": _float_or_none(row.get("cost_3y_per_10m_krw")),
            "total_expense_source_page": row.get("total_expense_source_page", ""),
            "synthetic_expense_source_page": row.get("synthetic_expense_source_page", ""),
            "cost_3y_source_page": row.get("cost_3y_source_page", ""),
        })
    return normalized


def row_key(row: dict) -> tuple[str, str]:
    return row["product_code"], row["class_code"]


def _values_match(left: Any, right: Any) -> bool:
    l_val = _float_or_none(left)
    r_val = _float_or_none(right)
    if l_val is None or r_val is None:
        return False
    return abs(float(l_val) - float(r_val)) <= 0.0001


def _infer_reference_cost_metric(parser: dict | None, reference: dict | None) -> str:
    """Infer which parser metric the workbook's generic cost value matches."""
    if not parser or not reference:
        return ""
    reference_value = reference.get("total_expense_ratio")
    for field in COST_METRIC_FIELDS:
        if _values_match(parser.get(field), reference_value):
            return field
    return ""


def compare_parser_to_reference(parser_rows: list[dict], reference_rows: list[dict]) -> list[dict]:
    parser_by_key = {row_key(row): row for row in parser_rows}
    reference_by_key = {row_key(row): row for row in reference_rows}
    all_keys = sorted(set(parser_by_key) | set(reference_by_key))
    results: list[dict] = []

    for key in all_keys:
        parser = parser_by_key.get(key)
        reference = reference_by_key.get(key)
        if parser and reference:
            notes = []
            if parser["account_type"] != reference["account_type"]:
                notes.append("account_type")
            if parser["channel"] != reference["channel"]:
                notes.append("channel")
            parser_has_cost = any(parser.get(field) is not None for field in COST_METRIC_FIELDS)
            r_ter = reference.get("total_expense_ratio")
            reference_metric = _infer_reference_cost_metric(parser, reference)
            if not parser_has_cost:
                notes.append("total_expense_ratio")
            elif r_ter is None:
                notes.append("reference_total_expense_ratio_missing")
            elif not reference_metric:
                notes.append("total_expense_ratio")
            elif (
                parser.get("total_expense_ratio") is not None
                and parser.get("synthetic_total_expense_ratio") is not None
                and not _values_match(parser.get("total_expense_ratio"), parser.get("synthetic_total_expense_ratio"))
            ):
                notes.append(f"reference_cost_metric={reference_metric}")
            if not notes:
                status = "MATCH"
            elif notes == ["reference_total_expense_ratio_missing"]:
                status = "REFERENCE_INCOMPLETE"
            elif all(note.startswith("reference_cost_metric=") for note in notes):
                status = "METRIC_SEMANTIC_MISMATCH"
            else:
                status = "FIELD_MISMATCH"
        elif parser:
            status = "EXTRA_IN_PARSER"
            notes = ["reference_missing"]
        else:
            status = "MISSING_IN_PARSER"
            notes = ["parser_missing"]

        source = parser or reference or {}
        results.append({
            "product_code": key[0],
            "class_code": key[1],
            "validation_status": status,
            "parser_account_type": parser.get("account_type") if parser else "",
            "reference_account_type": reference.get("account_type") if reference else "",
            "parser_channel": parser.get("channel") if parser else "",
            "reference_channel": reference.get("channel") if reference else "",
            "parser_total_expense_ratio": parser.get("total_expense_ratio") if parser else "",
            "parser_synthetic_total_expense_ratio": parser.get("synthetic_total_expense_ratio") if parser else "",
            "parser_cost_3y_per_10m_krw": parser.get("cost_3y_per_10m_krw") if parser else "",
            "reference_total_expense_ratio": reference.get("total_expense_ratio") if reference else "",
            "reference_cost_metric": _infer_reference_cost_metric(parser, reference),
            "reference_cost_3y_per_10m_krw": reference.get("cost_3y_per_10m_krw") if reference else "",
            "parser_source_file": parser.get("source_file") if parser else "",
            "reference_source_file": reference.get("source_file") if reference else "",
            "reference_status": reference.get("reference_status") if reference else "",
            "difference_note": " | ".join(notes),
            "class_label": source.get("class_label", ""),
        })
    return results


def canonical_rows(parser_rows: list[dict], validation_rows: list[dict]) -> list[dict]:
    validation_by_key = {(r["product_code"], r["class_code"]): r["validation_status"] for r in validation_rows}
    field_mismatch_by_key = {
        (r["product_code"], r["class_code"])
        for r in validation_rows
        if r["validation_status"] == "FIELD_MISMATCH"
    }
    rows = []
    for row in parser_rows:
        status = validation_by_key.get(row_key(row), "EXTRA_IN_PARSER")
        if row_key(row) in field_mismatch_by_key:
            continue
        rows.append({**row, "validation_status": status})
    return rows


def read_review_rows(path: Path | None) -> list[dict]:
    if path is None or not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = []
        for row in csv.DictReader(f):
            if not row.get("review_status"):
                continue
            # The previous P0 audit used a single total_expense_ratio world.
            # Do not auto-carry those overrides into the split metric schema.
            if "P0 PDF audit draft" in row.get("note", ""):
                continue
            rows.append(row)
        return rows


def validate_review_rows(rows: list[dict]) -> list[str]:
    errors = []
    for index, row in enumerate(rows, start=2):
        prefix = f"review row {index} {row.get('product_code')}/{row.get('class_code')}"
        status = row.get("review_status")
        field = row.get("field")
        if status not in REVIEW_STATUSES:
            errors.append(f"{prefix}: invalid review_status={status}")
        if field not in REVIEW_FIELDS:
            errors.append(f"{prefix}: invalid field={field}")
        if status == "BOTH_WRONG":
            if not row.get("reviewed_value"):
                errors.append(f"{prefix}: BOTH_WRONG requires reviewed_value")
            if not row.get("source_page"):
                errors.append(f"{prefix}: BOTH_WRONG requires source_page")
        if status == "SOURCE_CONFLICT":
            if not row.get("source_page"):
                errors.append(f"{prefix}: SOURCE_CONFLICT requires source_page")
            if not row.get("note"):
                errors.append(f"{prefix}: SOURCE_CONFLICT requires note")
        if status == "REFERENCE_CORRECT" and not row.get("reference_value") and not row.get("reviewed_value"):
            errors.append(f"{prefix}: REFERENCE_CORRECT requires reference_value or reviewed_value")
    return errors


def _review_value(row: dict) -> Any:
    status = row["review_status"]
    if status == "PARSER_CORRECT":
        return row.get("parser_value")
    if status == "REFERENCE_CORRECT":
        return row.get("reviewed_value") or row.get("reference_value")
    if status == "BOTH_WRONG":
        return row.get("reviewed_value")
    return None


def apply_review_overrides(
    canonical: list[dict],
    parser_rows: list[dict],
    validation_rows: list[dict],
    review_rows: list[dict],
) -> list[dict]:
    if not review_rows:
        return canonical

    canonical_by_key = {row_key(row): dict(row) for row in canonical}
    parser_by_key = {row_key(row): row for row in parser_rows}
    validation_by_key = {(row["product_code"], row["class_code"]): row for row in validation_rows}

    for review in review_rows:
        key = (review["product_code"], normalize_class_code(review["class_code"]))
        if review["review_status"] in {"AMBIGUOUS", "SOURCE_CONFLICT"}:
            canonical_by_key.pop(key, None)
            continue
        validation = validation_by_key.get(key, {})
        base = dict(canonical_by_key.get(key) or parser_by_key.get(key) or _base_row_from_validation(validation))
        if not base:
            continue
        field = review["field"]
        value = _review_value(review)
        if field in PRIMARY_REVIEW_FIELDS:
            value = _float_or_none(value)
        if value in (None, ""):
            continue
        base[field] = value
        base["validation_status"] = f"REVIEWED_{review['review_status']}"
        base["review_source_page"] = review.get("source_page", "")
        base["review_note"] = review.get("note", "")
        if validation:
            base["validation_status_before_review"] = validation["validation_status"]
        canonical_by_key[key] = base

    return [canonical_by_key[key] for key in sorted(canonical_by_key)]


def review_provenance_rows(
    canonical_before_review: list[dict],
    canonical_after_review: list[dict],
    review_rows: list[dict],
) -> list[dict]:
    """review.csv 판정이 canonical row에 어떤 영향을 줬는지 추적한다."""
    before_by_key = {row_key(row): row for row in canonical_before_review}
    after_by_key = {row_key(row): row for row in canonical_after_review}
    rows: list[dict] = []
    for review in review_rows:
        key = (review["product_code"], normalize_class_code(review["class_code"]))
        before = before_by_key.get(key)
        after = after_by_key.get(key)
        status = review["review_status"]
        if status in {"AMBIGUOUS", "SOURCE_CONFLICT"}:
            action = "REMOVED_FROM_CANONICAL" if before and not after else "EXCLUDED_FROM_CANONICAL"
        elif not before and after:
            action = "RESTORED_TO_CANONICAL"
        elif before and after:
            action = "UPDATED_EXISTING_CANONICAL"
        else:
            action = "NOT_APPLIED"
        field = review["field"]
        rows.append({
            "product_code": review["product_code"],
            "class_code": normalize_class_code(review["class_code"]),
            "field": field,
            "review_status": status,
            "canonical_action": action,
            "parser_value": review.get("parser_value", ""),
            "reference_value": review.get("reference_value", ""),
            "reviewed_value": review.get("reviewed_value", ""),
            "value_used": after.get(field, "") if after else "",
            "source_page": review.get("source_page", ""),
            "validation_status_after_review": after.get("validation_status", "") if after else "",
            "note": review.get("note", ""),
        })
    return rows


def _base_row_from_validation(row: dict) -> dict:
    if not row:
        return {}
    if row.get("validation_status") not in {"MISSING_IN_PARSER", "FIELD_MISMATCH"}:
        return {}
    ratio = _float_or_none(row.get("reference_total_expense_ratio"))
    if ratio is None:
        ratio = _float_or_none(row.get("parser_total_expense_ratio"))
    return {
        "product_code": row.get("product_code", ""),
        "class_code": row.get("class_code", ""),
        "account_type": row.get("reference_account_type") or row.get("parser_account_type"),
        "channel": row.get("reference_channel") or row.get("parser_channel"),
        "total_expense_ratio": ratio,
        "synthetic_total_expense_ratio": _float_or_none(row.get("parser_synthetic_total_expense_ratio")),
        "cost_3y_per_10m_krw": _float_or_none(row.get("reference_cost_3y_per_10m_krw"))
        or _float_or_none(row.get("parser_cost_3y_per_10m_krw")),
        "total_expense_source_page": "",
        "synthetic_expense_source_page": "",
        "cost_3y_source_page": "",
        "class_label": row.get("class_label", ""),
        "source_file": row.get("reference_source_file") or row.get("parser_source_file"),
        "parse_status": "review_override",
        "validation_status": row.get("validation_status", ""),
    }


def validate_canonical_rows(rows: list[dict]) -> list[str]:
    errors: list[str] = []
    seen = set()
    for row in rows:
        key = row_key(row)
        if key in seen:
            errors.append(f"duplicate key: {key}")
        seen.add(key)
        for field in ("product_code", "class_code", "account_type", "channel", "source_file"):
            if not row.get(field):
                errors.append(f"{key} missing {field}")
        if row.get("account_type") not in ALLOWED_ACCOUNT_TYPES:
            errors.append(f"{key} invalid account_type={row.get('account_type')}")
        if row.get("channel") not in ALLOWED_CHANNELS:
            errors.append(f"{key} invalid channel={row.get('channel')}")
        cost_values = []
        for field in COST_METRIC_FIELDS:
            value = _float_or_none(row.get(field))
            if value is not None:
                cost_values.append((field, value))
        if not cost_values:
            errors.append(f"{key} missing comparable annual cost metric")
        for field, value in cost_values:
            if not (0.01 <= float(value) <= 3.0):
                errors.append(f"{key} invalid {field}={value}")

    vip = {
        row["class_code"]: float(row.get("synthetic_total_expense_ratio") or row.get("total_expense_ratio"))
        for row in rows
        if row["product_code"] == "KR514X450008"
    }
    for code, expected in GOLDEN_VIP.items():
        actual = vip.get(code)
        if actual is None or abs(actual - expected) > 0.0001:
            errors.append(f"VIP golden mismatch {code}: expected {expected}, actual {actual}")
    return errors


def _comparison_kind(source_channel: str, target_channel: str) -> str:
    if source_channel in STANDARD_CHANNELS and target_channel in STANDARD_CHANNELS:
        return "STANDARD"
    return "CHANNEL_CONDITIONAL"


def choose_cost_metric(current: dict, target: dict) -> str | None:
    for field in COST_METRIC_FIELDS:
        if _float_or_none(current.get(field)) is not None and _float_or_none(target.get(field)) is not None:
            return field
    return None


def coverage_summary(rows: list[dict], validation_rows: list[dict]) -> dict:
    by_fund_account: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        by_fund_account[(row["product_code"], row["account_type"])].append(row)

    comparable_groups = []
    pairs = []
    excluded_no_common_metric = 0
    for (product_code, account_type), group in sorted(by_fund_account.items()):
        group_has_pair = False
        if len(group) < 2:
            continue
        for current in group:
            for target in group:
                if row_key(current) == row_key(target):
                    continue
                metric = choose_cost_metric(current, target)
                if metric is None:
                    excluded_no_common_metric += 1
                    continue
                current_value = float(current[metric])
                target_value = float(target[metric])
                if target_value < current_value:
                    group_has_pair = True
                    pairs.append({
                        "product_code": product_code,
                        "account_type": account_type,
                        "current_class_code": current["class_code"],
                        "current_channel": current["channel"],
                        "comparison_metric": metric,
                        "current_value": current_value,
                        "current_total_expense_ratio": current.get("total_expense_ratio"),
                        "current_synthetic_total_expense_ratio": current.get("synthetic_total_expense_ratio"),
                        "current_cost_3y_per_10m_krw": current.get("cost_3y_per_10m_krw"),
                        "target_class_code": target["class_code"],
                        "target_channel": target["channel"],
                        "target_value": target_value,
                        "target_total_expense_ratio": target.get("total_expense_ratio"),
                        "target_synthetic_total_expense_ratio": target.get("synthetic_total_expense_ratio"),
                        "target_cost_3y_per_10m_krw": target.get("cost_3y_per_10m_krw"),
                        "difference_pct_point": round(current_value - target_value, 6),
                        "comparison_kind": _comparison_kind(current["channel"], target["channel"]),
                    })
        if group_has_pair:
            comparable_groups.append((product_code, account_type))

    validation_counts = Counter(row["validation_status"] for row in validation_rows)
    pair_counts = Counter(row["comparison_kind"] for row in pairs)
    metric_counts = Counter(row["comparison_metric"] for row in pairs)
    pair_kind_metric_counts = Counter((row["comparison_kind"], row["comparison_metric"]) for row in pairs)
    account_counts = Counter(account_type for _, account_type in comparable_groups)
    return {
        "canonical_fund_count": len({row["product_code"] for row in rows}),
        "canonical_row_count": len(rows),
        "validation_counts": dict(sorted(validation_counts.items())),
        "comparable_fund_account_group_count": len(comparable_groups),
        "comparable_fund_count": len({product_code for product_code, _ in comparable_groups}),
        "comparable_by_account_type": dict(sorted(account_counts.items())),
        "lower_cost_pair_count": len(pairs),
        "lower_cost_pair_by_kind": dict(sorted(pair_counts.items())),
        "lower_cost_pair_by_metric": dict(sorted(metric_counts.items())),
        "lower_cost_pair_by_kind_and_metric": {
            f"{kind}:{metric}": count
            for (kind, metric), count in sorted(pair_kind_metric_counts.items())
        },
        "excluded_pair_no_common_metric_count": excluded_no_common_metric,
        "lower_cost_pairs": pairs,
    }


def build_audit_rows(
    validation_rows: list[dict],
    parser_rows: list[dict],
    all_reference_rows: list[dict],
    *,
    sample_size: int = 20,
) -> list[dict]:
    parser_products = {row["product_code"] for row in parser_rows}
    parser_keys = {row_key(row) for row in parser_rows}
    all_reference_by_key = {row_key(row): row for row in all_reference_rows}
    sample_counts: Counter[str] = Counter()
    audit_rows: list[dict] = []

    for row in validation_rows:
        status = row["validation_status"]
        if status not in {
            "FIELD_MISMATCH",
            "METRIC_SEMANTIC_MISMATCH",
            "MISSING_IN_PARSER",
            "EXTRA_IN_PARSER",
            "REFERENCE_INCOMPLETE",
        }:
            continue

        key = (row["product_code"], row["class_code"])
        cause = "needs_manual_review"
        action = "PDF 원문 대조 필요"

        if status == "FIELD_MISMATCH":
            note = row["difference_note"]
            if "total_expense_ratio" in note:
                cause = "same_metric_mismatch"
                action = "동일 비용 metric 기준 값인지 원문 재확인"
            elif "account_type" in note:
                cause = "account_type_mismatch"
                action = "클래스 정의 원문에서 계좌유형 재확인"
            elif "channel" in note:
                cause = "channel_mismatch"
                action = "클래스 정의 원문에서 판매채널 재확인"
        elif status == "METRIC_SEMANTIC_MISMATCH":
            cause = "metric_semantic_mismatch"
            action = "서로 다른 비용 metric 값으로 확인되어 둘 다 보존"
        elif status == "REFERENCE_INCOMPLETE":
            cause = "reference_missing_expense_ratio"
            action = "파서값은 유지 후보, 엑셀 비용값 공란 원인 확인"
        elif status == "MISSING_IN_PARSER":
            ref_ratio = row["reference_total_expense_ratio"]
            if not ref_ratio:
                cause = "reference_missing_expense_ratio"
                action = "엑셀 기준값도 불완전하므로 Cost Guard 후보 제외 유지"
            elif row["product_code"] not in parser_products:
                cause = "parser_no_clean_fund"
                action = "해당 펀드가 dirty/review 상태인지 확인"
            else:
                cause = "parser_clean_fund_missing_class"
                action = "같은 펀드 내 특정 클래스 누락 원문 대조"
        elif status == "EXTRA_IN_PARSER":
            ref = all_reference_by_key.get(key)
            if ref is None:
                cause = "not_found_in_reference_workbook"
                action = "파서 오탐인지 엑셀 누락인지 원문 샘플 대조"
            elif ref.get("cost_guard_usable") != "Y":
                cause = "reference_not_cost_guard_usable"
                action = "엑셀에서 CostGuard 제외한 이유 확인"
            elif ref.get("account_type") not in ALLOWED_ACCOUNT_TYPES:
                cause = "reference_scope_excluded"
                action = "일반/기타 클래스 여부 확인"
            elif key not in parser_keys:
                cause = "normalization_or_key_mismatch"
                action = "클래스 코드 normalize 기준 확인"

        needs_pdf_review = status in {"FIELD_MISMATCH", "EXTRA_IN_PARSER"} or (
            status == "MISSING_IN_PARSER"
            and cause in {"parser_clean_fund_missing_class", "parser_no_clean_fund"}
        )
        sample_rank = ""
        if needs_pdf_review and sample_counts[status] < sample_size:
            sample_counts[status] += 1
            sample_rank = str(sample_counts[status])

        audit_rows.append({
            **row,
            "audit_cause": cause,
            "recommended_action": action,
            "needs_pdf_review": "Y" if needs_pdf_review else "N",
            "sample_rank": sample_rank,
        })
    return audit_rows


def audit_summary(audit_rows: list[dict]) -> dict:
    by_status = Counter(row["validation_status"] for row in audit_rows)
    by_cause = Counter(row["audit_cause"] for row in audit_rows)
    by_status_cause: dict[str, dict[str, int]] = {}
    for row in audit_rows:
        status = row["validation_status"]
        by_status_cause.setdefault(status, {})
        by_status_cause[status][row["audit_cause"]] = by_status_cause[status].get(row["audit_cause"], 0) + 1
    sample_rows = [
        {
            "product_code": row["product_code"],
            "class_code": row["class_code"],
            "validation_status": row["validation_status"],
            "audit_cause": row["audit_cause"],
            "recommended_action": row["recommended_action"],
        }
        for row in audit_rows
        if row["sample_rank"]
    ]
    return {
        "audit_row_count": len(audit_rows),
        "audit_by_status": dict(sorted(by_status.items())),
        "audit_by_cause": dict(sorted(by_cause.items())),
        "audit_by_status_cause": {
            status: dict(sorted(causes.items()))
            for status, causes in sorted(by_status_cause.items())
        },
        "pdf_review_sample_count": len(sample_rows),
        "pdf_review_samples": sample_rows,
    }


def review_template_rows(audit_rows: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for row in audit_rows:
        cause = row["audit_cause"]
        if cause not in {
            "same_metric_mismatch",
            "channel_mismatch",
            "parser_clean_fund_missing_class",
            "parser_no_clean_fund",
        }:
            continue
        fields: list[tuple[str, Any, Any]] = []
        if row["validation_status"] == "FIELD_MISMATCH":
            notes = set(filter(None, row["difference_note"].split(" | ")))
            if "total_expense_ratio" in notes or "reference_total_expense_ratio_missing" in notes:
                fields.append(("total_expense_ratio", row["parser_total_expense_ratio"], row["reference_total_expense_ratio"]))
            if "synthetic_total_expense_ratio" in notes:
                fields.append((
                    "synthetic_total_expense_ratio",
                    row["parser_synthetic_total_expense_ratio"],
                    row["reference_total_expense_ratio"],
                ))
            if "channel" in notes:
                fields.append(("channel", row["parser_channel"], row["reference_channel"]))
            if "account_type" in notes:
                fields.append(("account_type", row["parser_account_type"], row["reference_account_type"]))
        else:
            fields.append(("total_expense_ratio", row["parser_total_expense_ratio"], row["reference_total_expense_ratio"]))

        for field, parser_value, reference_value in fields:
            rows.append({
                "product_code": row["product_code"],
                "class_code": row["class_code"],
                "field": field,
                "parser_value": parser_value,
                "reference_value": reference_value,
                "review_status": "",
                "reviewed_value": "",
                "source_page": "",
                "note": f"{row['validation_status']} / {cause} / {row['recommended_action']}",
            })
    return rows


def p0_review_template_rows(audit_rows: list[dict]) -> list[dict]:
    """Cost Guard v1 freeze 전에 반드시 판정할 P0 행만 반환한다."""
    p0_causes = {"same_metric_mismatch", "channel_mismatch"}
    return [
        row
        for row in review_template_rows(audit_rows)
        if any(cause in (row.get("note") or "") for cause in p0_causes)
    ]


def _review_key(row: dict) -> tuple[str, str, str]:
    return row.get("product_code", ""), normalize_class_code(row.get("class_code")), row.get("field", "")


def unresolved_p0_review_rows(p0_rows: list[dict], review_rows: list[dict]) -> list[dict]:
    """P0 template 중 아직 review.csv에 판정이 없는 field row를 반환한다."""
    reviewed = {_review_key(row) for row in review_rows if row.get("review_status")}
    return [row for row in p0_rows if _review_key(row) not in reviewed]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_manifest(
    coverage: dict,
    *,
    canonical_path: Path,
    p0_case_count: int,
    p0_template: list[dict],
    unresolved_p0_rows: list[dict],
    review_rows: list[dict],
    review_errors: list[str],
    validation_errors: list[str],
) -> dict:
    is_frozen = not unresolved_p0_rows and not review_errors and not validation_errors
    return {
        "dataset_version": "cost_guard_v1" if is_frozen else "cost_guard_provisional",
        "dataset_status": "FROZEN_V1" if is_frozen else "PROVISIONAL",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "canonical_sha256": sha256_file(canonical_path),
        "canonical_row_count": coverage["canonical_row_count"],
        "canonical_fund_count": coverage["canonical_fund_count"],
        "standard_pair_count": coverage["lower_cost_pair_by_kind"].get("STANDARD", 0),
        "channel_conditional_pair_count": coverage["lower_cost_pair_by_kind"].get("CHANNEL_CONDITIONAL", 0),
        "review_override_count": len(review_rows),
        "p0_review_required_case_count": p0_case_count,
        "p0_review_required_field_row_count": len(p0_template),
        "p0_review_unresolved_field_row_count": len(unresolved_p0_rows),
    }


def _compact_snippet(text: str, limit: int = 900) -> str:
    compact = re.sub(r"\s+", " ", text).strip()
    return compact[:limit]


def _snippets_for_class(pdf_root: Path, product_code: str, class_code: str, *, max_snippets: int = 3) -> list[str]:
    pdf_path = pdf_root / product_code / f"R2_{product_code}.pdf"
    if not pdf_path.exists():
        return []
    pattern = re.compile(rf"(?<![A-Za-z0-9\-]){re.escape(class_code)}(?![A-Za-z0-9\-])", re.IGNORECASE)
    snippets: list[str] = []
    with pymupdf.open(str(pdf_path)) as doc:
        for page_index, page in enumerate(doc, start=1):
            text = page.get_text() or ""
            for match in pattern.finditer(text):
                start = max(0, match.start() - 450)
                end = min(len(text), match.end() + 450)
                snippets.append(f"p.{page_index}: {_compact_snippet(text[start:end])}")
                if len(snippets) >= max_snippets:
                    return snippets
    return snippets


def write_pdf_review_pack(path: Path, audit_rows: list[dict], pdf_root: Path) -> None:
    priority_causes = {
        "same_metric_mismatch",
        "channel_mismatch",
        "parser_clean_fund_missing_class",
        "parser_no_clean_fund",
        "not_found_in_reference_workbook",
    }
    rows = [
        row for row in audit_rows
        if row["audit_cause"] in priority_causes and row["needs_pdf_review"] == "Y"
    ]
    rows.sort(key=lambda row: (
        0 if row["audit_cause"] in {"same_metric_mismatch", "channel_mismatch"} else 1,
        row["product_code"],
        row["class_code"],
    ))

    lines = [
        "# Fund Class Pension PDF Review Pack",
        "",
        "이 파일은 원문 PDF 대조가 필요한 audit 행의 자동 스니펫입니다.",
        "스니펫은 최종 판정이 아니라 검수 시작점으로만 사용합니다.",
        "",
    ]
    for row in rows:
        lines.extend([
            f"## {row['product_code']} / {row['class_code']} / {row['audit_cause']}",
            "",
            f"- validation_status: {row['validation_status']}",
            (
                f"- parser: {row['parser_account_type']} / {row['parser_channel']} / "
                f"total={row['parser_total_expense_ratio']} / "
                f"synthetic={row.get('parser_synthetic_total_expense_ratio', '')}"
            ),
            (
                f"- reference: {row['reference_account_type']} / {row['reference_channel']} / "
                f"cost={row['reference_total_expense_ratio']} / "
                f"inferred_metric={row.get('reference_cost_metric', '')}"
            ),
            f"- recommended_action: {row['recommended_action']}",
            "",
        ])
        snippets = _snippets_for_class(pdf_root, row["product_code"], row["class_code"])
        if snippets:
            for snippet in snippets:
                lines.extend(["```text", snippet, "```", ""])
        else:
            lines.extend(["```text", "원문에서 class_code 직접 매칭 스니펫을 찾지 못했습니다.", "```", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf-root", type=Path, default=DEFAULT_PDF_ROOT)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--review-csv",
        type=Path,
        default=None,
        help="사람이 원문 대조 후 작성한 review override CSV",
    )
    args = parser.parse_args()

    parser_rows, dirty_funds = parser_rows_from_pdf_root(args.pdf_root)
    parser_rows = normalize_parser_rows(parser_rows)
    reference_rows = read_reference_rows(args.xlsx)
    all_reference_rows = read_reference_rows(args.xlsx, cost_guard_usable_only=False)
    validation_rows = compare_parser_to_reference(parser_rows, reference_rows)
    audit_rows = build_audit_rows(validation_rows, parser_rows, all_reference_rows)
    review_path = args.review_csv or args.output_dir / "fund_class_pension_review.csv"
    review_rows = read_review_rows(review_path)
    review_errors = validate_review_rows(review_rows)
    canonical_before_review = canonical_rows(parser_rows, validation_rows)
    canonical = apply_review_overrides(canonical_before_review, parser_rows, validation_rows, review_rows)
    provenance_rows = review_provenance_rows(canonical_before_review, canonical, review_rows)
    errors = validate_canonical_rows(canonical)
    coverage = coverage_summary(canonical, validation_rows)
    audit = audit_summary(audit_rows)
    coverage["automatic_clean_fund_count"] = len({row["product_code"] for row in parser_rows})
    coverage["automatic_clean_row_count"] = len(parser_rows)
    coverage["dirty_fund_count"] = len(dirty_funds)
    coverage["dirty_funds"] = dirty_funds
    coverage["audit"] = audit
    p0_template = p0_review_template_rows(audit_rows)
    unresolved_p0 = unresolved_p0_review_rows(p0_template, review_rows)
    p0_case_count = sum(
        1
        for row in audit_rows
        if row.get("audit_cause") in {"same_metric_mismatch", "channel_mismatch"}
    )
    coverage["p0_review_required_case_count"] = p0_case_count
    coverage["p0_review_required_field_row_count"] = len(p0_template)
    coverage["p0_review_unresolved_field_row_count"] = len(unresolved_p0)
    coverage["review_override_path"] = str(review_path)
    coverage["review_override_count"] = len(review_rows)
    coverage["canonical_row_count_before_review"] = len(canonical_before_review)
    coverage["canonical_row_count_delta_from_review"] = len(canonical) - len(canonical_before_review)
    coverage["review_provenance_action_counts"] = dict(sorted(Counter(row["canonical_action"] for row in provenance_rows).items()))
    coverage["review_validation_errors"] = review_errors
    coverage["validation_errors"] = errors

    canonical_path = args.output_dir / "fund_class_pension.csv"
    write_csv(
        canonical_path,
        canonical,
        [
            "product_code",
            "class_code",
            "account_type",
            "channel",
            "total_expense_ratio",
            "synthetic_total_expense_ratio",
            "cost_3y_per_10m_krw",
            "total_expense_source_page",
            "synthetic_expense_source_page",
            "cost_3y_source_page",
            "class_label",
            "source_file",
            "parse_status",
            "validation_status",
            "validation_status_before_review",
            "review_source_page",
            "review_note",
        ],
    )
    write_csv(
        args.output_dir / "fund_class_pension_validation.csv",
        validation_rows,
        [
            "product_code",
            "class_code",
            "validation_status",
            "parser_account_type",
            "reference_account_type",
            "parser_channel",
            "reference_channel",
            "parser_total_expense_ratio",
            "parser_synthetic_total_expense_ratio",
            "parser_cost_3y_per_10m_krw",
            "reference_total_expense_ratio",
            "reference_cost_metric",
            "reference_cost_3y_per_10m_krw",
            "parser_source_file",
            "reference_source_file",
            "reference_status",
            "difference_note",
            "class_label",
        ],
    )
    write_csv(
        args.output_dir / "fund_class_pension_audit.csv",
        audit_rows,
        [
            "product_code",
            "class_code",
            "validation_status",
            "audit_cause",
            "recommended_action",
            "needs_pdf_review",
            "sample_rank",
            "parser_account_type",
            "reference_account_type",
            "parser_channel",
            "reference_channel",
            "parser_total_expense_ratio",
            "parser_synthetic_total_expense_ratio",
            "parser_cost_3y_per_10m_krw",
            "reference_total_expense_ratio",
            "reference_cost_metric",
            "reference_cost_3y_per_10m_krw",
            "parser_source_file",
            "reference_source_file",
            "reference_status",
            "difference_note",
            "class_label",
        ],
    )
    write_csv(
        args.output_dir / "fund_class_pension_review_template.csv",
        review_template_rows(audit_rows),
        REVIEW_FIELDNAMES,
    )
    write_csv(
        args.output_dir / "fund_class_pension_p0_review_template.csv",
        p0_template,
        REVIEW_FIELDNAMES,
    )
    write_csv(
        args.output_dir / "fund_class_pension_review_provenance.csv",
        provenance_rows,
        [
            "product_code",
            "class_code",
            "field",
            "review_status",
            "canonical_action",
            "parser_value",
            "reference_value",
            "reviewed_value",
            "value_used",
            "source_page",
            "validation_status_after_review",
            "note",
        ],
    )
    coverage_path = args.output_dir / "fund_class_pension_coverage.json"
    coverage_path.write_text(json.dumps(coverage, ensure_ascii=False, indent=2), encoding="utf-8")
    manifest = build_manifest(
        coverage,
        canonical_path=canonical_path,
        p0_case_count=p0_case_count,
        p0_template=p0_template,
        unresolved_p0_rows=unresolved_p0,
        review_rows=review_rows,
        review_errors=review_errors,
        validation_errors=errors,
    )
    (args.output_dir / "fund_class_pension_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_pdf_review_pack(args.output_dir / "fund_class_pension_pdf_review.md", audit_rows, args.pdf_root)

    print(f"자동 파싱 clean: {coverage['automatic_clean_fund_count']}개 펀드 / {coverage['automatic_clean_row_count']}건")
    print(f"canonical: {coverage['canonical_fund_count']}개 펀드 / {coverage['canonical_row_count']}건")
    print(f"validation: {coverage['validation_counts']}")
    print(f"audit: {audit['audit_by_status_cause']}")
    print(f"review overrides: {len(review_rows)}")
    print(
        f"P0 review required: {p0_case_count} cases / {len(p0_template)} field rows"
        f" / unresolved {len(unresolved_p0)}"
    )
    print(f"dataset status: {manifest['dataset_status']} ({manifest['dataset_version']})")
    print(f"lower-cost pairs: {coverage['lower_cost_pair_count']} {coverage['lower_cost_pair_by_kind']}")
    if review_errors:
        raise SystemExit("review_validation_errors: " + "; ".join(review_errors[:5]))
    if errors:
        raise SystemExit("validation_errors: " + "; ".join(errors[:5]))


if __name__ == "__main__":
    main()
