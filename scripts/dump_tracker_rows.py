import sys
import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서.xlsm"
OUT = r"C:\Users\kevin\pension-agent\prospectus_check\_tracker_12_17.txt"
codes = sys.argv[1:]

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱"]

header = [c.value for c in ws[5]]

with io.open(OUT, "a", encoding="utf-8") as f:
    for code in codes:
        f.write(f"\n===== {code} =====\n")
        for row in ws.iter_rows(min_row=6):
            if row[2].value == code:
                for h, cell in zip(header, row):
                    f.write(f"{h}: {cell.value}\n")
                f.write("---\n")
