import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_recheck_sheet.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_재검수_추가수정필요"]

with io.open(OUT, "w", encoding="utf-8") as f:
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if any(v is not None and str(v).strip() != "" for v in vals):
            f.write(f"row{r}: {vals}\n\n")
