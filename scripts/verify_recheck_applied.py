import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_verify_recheck_applied.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

TARGETS = [("KR515302022M", "C1"), ("KR515302022M", "Ce"), ("KR5153420318", "C-e"),
           ("KR510902511M", "C"), ("KR510902511M", "C-e"), ("KR516702010M", "C")]

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"max_row(현재): {ws.max_row}\n\n")
    found = {t: False for t in TARGETS}
    for r in range(6, ws.max_row + 1):
        code = ws.cell(row=r, column=3).value
        cls = ws.cell(row=r, column=12).value
        if (code, cls) in TARGETS:
            found[(code, cls)] = True
            f.write(f"row{r}: {code} ({cls}) 1y={ws.cell(row=r,column=17).value} "
                    f"3y={ws.cell(row=r,column=18).value} since={ws.cell(row=r,column=19).value} "
                    f"최초설정일={ws.cell(row=r,column=35).value} 검수상태={ws.cell(row=r,column=38).value}\n")
    f.write("\n--- 못 찾은 대상(삭제됐어야 하는 것 제외하고 확인) ---\n")
    for t, ok in found.items():
        if not ok:
            f.write(f"{t}: 없음\n")
