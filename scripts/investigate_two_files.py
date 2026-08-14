import io
import openpyxl

OUT = r"C:\Users\kevin\pension-agent\docs_check\_investigate_two_files.txt"

paths = {
    "docs 수정.xlsm": r"C:\Users\kevin\Downloads\docs 수정.xlsm",
    "[파싱]renamed폴더.xlsm": r"C:\Users\kevin\Downloads\[파싱]renamed폴더.xlsm",
}

with io.open(OUT, "w", encoding="utf-8") as f:
    for label, path in paths.items():
        f.write(f"===== {label} =====\n")
        wb = openpyxl.load_workbook(path, data_only=True)
        f.write(f"시트 목록: {wb.sheetnames}\n\n")
        for sheet_name in wb.sheetnames:
            if "파싱" in sheet_name or "결과" in sheet_name:
                ws = wb[sheet_name]
                f.write(f"--- 시트 '{sheet_name}' (dims={ws.dimensions}) ---\n")
                # find doc36_chunk04 row
                for row in ws.iter_rows(min_row=1):
                    for cell in row:
                        if cell.value == "doc36_chunk04":
                            r = cell.row
                            f.write(f"doc36_chunk04 found at row {r}\n")
                            for c in ws[r]:
                                if c.value:
                                    f.write(f"  col{c.column}: {repr(c.value)[:300]}\n")
                            break
                f.write("\n")
        f.write("\n")
