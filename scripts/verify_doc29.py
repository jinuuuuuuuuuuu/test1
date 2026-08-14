import io
import openpyxl

TRACKER_PATH = r"C:\Users\kevin\Downloads\docs 수정.xlsm"
SRC_PATH = r"C:\Users\kevin\pension-agent\data\raw\docs\docs_renamed\doc29.xlsx"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_doc29_diff.txt"

wb_src = openpyxl.load_workbook(SRC_PATH, data_only=True)
ws_src = wb_src["FAQ_100"]

src_rows = {}
for row in ws_src.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    idx = int(row[0])
    src_rows[idx] = {
        "카테고리": row[1],
        "대표질문": row[2],
        "유사질문표현": row[3],
        "표준답변": row[4],
        "AI분기포인트": row[5],
        "근거구분": row[6],
        "근거ID": row[7],
    }

wb_t = openpyxl.load_workbook(TRACKER_PATH, data_only=True)
ws_t = wb_t["파싱 결과"]

diffs = []
count = 0
for row in ws_t.iter_rows(min_row=6, values_only=True):
    if row[0] is None:
        continue
    try:
        num = int(float(row[0]))
    except (TypeError, ValueError):
        continue
    if num != 29:
        continue
    count += 1
    chunk_id = row[7]
    chunk_text = row[8] or ""
    # chunk_id like doc29_chunk07 -> row index = chunk number + 0 (chunk01 -> FAQ row ID 1)
    try:
        n = int(chunk_id.split("chunk")[1])
    except Exception:
        diffs.append(f"{chunk_id}: cannot parse chunk number")
        continue
    src = src_rows.get(n)
    if src is None:
        diffs.append(f"{chunk_id}: no matching FAQ_100 row ID={n}")
        continue
    problems = []
    for field in ["대표질문", "유사질문표현", "표준답변", "AI분기포인트", "근거ID"]:
        val = src[field]
        if val is None:
            continue
        val_str = str(val).strip()
        if val_str not in chunk_text:
            problems.append(f"[{field}] MISMATCH\n  SRC: {val_str}\n  CHUNK contains: {'있음' if field in chunk_text else '없음'}")
    if problems:
        diffs.append(f"{chunk_id} (FAQ ID={n}):\n" + "\n".join(problems))

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"총 tracker doc29 청크 수: {count}, 원본 FAQ 행 수: {len(src_rows)}\n\n")
    if diffs:
        f.write(f"불일치 {len(diffs)}건:\n\n")
        f.write("\n\n".join(diffs))
    else:
        f.write("불일치 없음 - 전체 일치\n")
