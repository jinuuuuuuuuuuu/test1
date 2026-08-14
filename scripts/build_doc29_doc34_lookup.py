"""doc29(디폴트옵션 FAQ), doc34(실물이전 25코드)를 파싱가이드(sheet2 '예외 자료 처리')가
지정한 lookup 테이블 형식으로 원본 xlsx에서 직접 재구성한다.

- doc29 저장필드: question / answer / keywords (+ 추적용으로 category, similar_questions, source_ids 추가)
- doc34 저장필드: code / reason / description / action

원본을 직접 openpyxl로 읽어서 만든다 (수기 전사 시 발생할 수 있는 오류를 피하기 위함).
"""

import json
from pathlib import Path

import openpyxl

DOCS_DIR = Path(r"C:\Users\kevin\pension-agent\data\raw\docs\docs_renamed")
OUT_DIR = Path(r"C:\Users\kevin\pension-agent\data\processed")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def build_doc29_lookup():
    wb = openpyxl.load_workbook(DOCS_DIR / "doc29.xlsx", data_only=True)
    faq_sheet = wb["FAQ_100"]
    sources_sheet = wb["Sources"]

    source_map = {}
    for row in sources_sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sid, desc, url, note = (list(row) + [None] * 4)[:4]
        source_map[str(sid).strip()] = {"desc": desc, "url": url, "note": note}

    entries = []
    for row in faq_sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rid, category, question, similar, answer, branch_points, source_type, source_ids = (
            list(row) + [None] * 8
        )[:8]

        similar_list = [s.strip() for s in (similar or "").split("/") if s.strip()]
        branch_list = [s.strip() for s in (branch_points or "").split(",") if s.strip()]
        source_id_list = [s.strip() for s in (source_ids or "").split(",") if s.strip()]

        entries.append(
            {
                "id": int(rid),
                "category": category,
                "question": question,
                "similar_questions": similar_list,
                "answer": answer,
                "keywords": branch_list,
                "source_type": source_type,
                "source_ids": source_id_list,
            }
        )

    out = {"entries": entries, "sources": source_map}
    out_path = OUT_DIR / "doc29_default_option_qa_lookup.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"doc29: {len(entries)} entries -> {out_path}")
    return len(entries)


def build_doc34_lookup():
    wb = openpyxl.load_workbook(DOCS_DIR / "doc34.xlsx", data_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    # 액션 안내 — 원문(설명)에는 없는, 재원/처리방향에 대한 실무 요약을 코드별로 덧붙인다.
    # (in_kind_transfer.py 의 TRANSFER_BLOCK_CODES와 코드 체계를 동일하게 맞춤)
    action_by_code = {
        "01": "소규모 펀드 임의해지 대상이면 현금화(매도) 후 이관해야 함 — 실물 그대로는 이전 불가",
        "02": "언번들계약이면 자산관리기관 기준으로 별도 확인 필요 — 대표 언번들 플랜 목록 참고",
        "03": "사모펀드는 실물이전 불가 — 매도 후 현금이전 검토",
        "04": "MMF는 실물이전 불가 — 매도 후 현금이전 검토",
        "05": "환매수수료 부과 상품은 수수료 발생 여부 확인 후 진행",
        "06": "이관은 가능하나 수관기관의 상품제공수수료 협약 여부에 따라 수관 불가할 수 있음 — 수관기관 확인 필요 (directional)",
        "07": "사전조회 단계면 해당 상품만 제외, 이전접수 단계면 계좌 전체 이전접수 불가 — 운용지시 완료 후 재신청",
        "08": "압류·질권 해제 전까지 이전 불가",
        "09": "만기매칭형 펀드는 실물이전 불가 — 만기 도래 후 처리",
        "10": "지분증권/리츠는 권리 처리 문제로 실물이전 불가 — 매도 후 현금이전",
        "11": "RP는 실물이전 불가 — 매도 후 현금이전",
        "12": "발행어음은 실물이전 불가 — 매도 후 현금이전",
        "13": "금리연동형 보험은 실물이전 불가 대상으로 확인 필요",
        "14": "실적배당형 보험은 불가, 이율보증형(GIC)만 예외적으로 실물이전 가능",
        "15": "원금비보장 파생결합사채는 불가, 원리금보장 ELB·DLB만 가능",
        "16": "수관기관에 DC 규약 체결 후 재신청",
        "17": "이관/수관 상대기관의 상품라인업 포함 여부를 개별 확인 필요 (directional)",
        "18": "저축은행예금은 예금자보호한도(1억원, 저축은행별 합산) 이내로 조정 후 이전",
        "19": "수관기관 자사 원리금보장상품은 수관 불가 — 타 상품으로 대체 필요",
        "20": "재원 구분 여부와 무관하게 이관 자체는 가능 — 수관기관의 수용 가능 여부만 확인 (directional, 당사는 구분/미구분 모두 수관 가능)",
        "21": "이미 만기 도래한 상품은 실물이 아닌 상환금으로 처리 — 만기 처리 절차를 따름",
        "22": "환매불가 기간 종료 후 재신청",
        "23": "디폴트옵션 상품은 실물이전 자체가 불가 — 매도 후 현금이전만 가능",
        "24": "맥쿼리인프라 등 상장투자회사는 실물이전 불가 — 매도 후 현금이전",
        "25": "수관기관과 상품 제공기관 간 상품협약(위탁계약) 체결 후 재신청",
        "99": "개별 사유(소규모펀드/합병전 매수상품/가입자 미분리 매수상품 등)를 확인 후 담당자와 개별 협의",
    }

    entries = []
    # 실제 코드표는 B/C열, 7행부터 시작 (1~6행은 제목/설명/헤더)
    for row in sheet.iter_rows(min_row=7, min_col=2, max_col=3, values_only=True):
        code_raw, desc = (list(row) + [None] * 2)[:2]
        if code_raw is None:
            continue
        code_raw = str(code_raw).strip()
        # "01. 소규모 펀드 임의해지" 형태에서 코드와 사유명 분리
        parts = code_raw.split(".", 1)
        code = parts[0].strip().zfill(2) if parts[0].strip().isdigit() else parts[0].strip()
        reason = parts[1].strip() if len(parts) > 1 else code_raw

        entries.append(
            {
                "code": code,
                "reason": reason,
                "description": (desc or "").strip(),
                "action": action_by_code.get(code, "담당자 확인 필요"),
            }
        )

    out = {"entries": entries}
    out_path = OUT_DIR / "doc34_in_kind_transfer_code_lookup.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"doc34: {len(entries)} entries -> {out_path}")
    return len(entries)


if __name__ == "__main__":
    n29 = build_doc29_lookup()
    n34 = build_doc34_lookup()
    print(f"Done. doc29={n29} entries, doc34={n34} entries")
