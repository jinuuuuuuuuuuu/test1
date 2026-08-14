import openpyxl

PATH = r"C:\Users\kevin\Downloads\docs 수정.xlsm"

wb = openpyxl.load_workbook(PATH, keep_vba=True)
ws = wb["파싱 결과"]


def find_row(chunk_id):
    for row in ws.iter_rows(min_row=6):
        if row[7].value == chunk_id:
            return row
    raise ValueError(f"chunk_id not found: {chunk_id}")


# doc41_chunk03의 표구조화텍스트 필드도 동일하게 보완 (chunk_text만 고치고 빠뜨렸던 부분)
r = find_row("doc41_chunk03")
before = r[10].value
r[10].value = r[10].value.replace(
    "5,500만 원 초과 (종합소득 4,500만 원 초과)\t13.2%\t\t118만 8천 원",
    "5,500만 원 초과 (종합소득 4,500만 원 초과)\t13.2%\t연 900만 원 (연금저축 단독 600만원)\t118만 8천 원",
)
assert r[10].value != before, "doc41_chunk03 표구조화텍스트 replace failed"

# doc39는 전체 검증 완료(원문과 100% 일치)했는데 일괄전환 목록에서 누락됐었음
# doc20/36/38/40/41의 나머지 청크들(직접 수정한 것 제외)도 원문 대조 완료 -> 검수완료 전환
REMAINING_CLEAN_DOCS = [39]
count = 0
for row in ws.iter_rows(min_row=6):
    fnum = row[0].value
    if fnum is None:
        continue
    try:
        num = int(float(fnum))
    except (TypeError, ValueError):
        continue
    if num in REMAINING_CLEAN_DOCS and row[12].value == "검수전":
        row[12].value = "검수완료"
        count += 1

# 개별 처리했던 문서(20,36,38,40,41)의 '나머지' 청크들 -> 이미 원문 대조 완료된 것들
REMAINING_CLEAN_CHUNKS = [
    "doc20_chunk01",
    "doc36_chunk01", "doc36_chunk02", "doc36_chunk03", "doc36_chunk05",
    "doc38_chunk01", "doc38_chunk03", "doc38_chunk04",
    "doc40_chunk02", "doc40_chunk03", "doc40_chunk04", "doc40_chunk05",
    "doc41_chunk01", "doc41_chunk02",
]
for cid in REMAINING_CLEAN_CHUNKS:
    row = find_row(cid)
    if row[12].value == "검수전":
        row[12].value = "검수완료"
        count += 1

wb.save(PATH)
print(f"doc41_chunk03 표구조화텍스트 수정 완료. 추가 상태전환: {count}개")
