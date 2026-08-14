import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (1).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_null_scale.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

header_row = ws[1]
COLS = {16: "수익률_최근1년", 17: "수익률_최근3년", 18: "수익률_설정일이후"}

null_rows = []
total_rows = 0
for row in ws.iter_rows(min_row=2):
    if row[2].value is None:  # 상품코드 컬럼 비어있으면 skip
        continue
    total_rows += 1
    vals = {idx: row[idx].value for idx in COLS}
    is_null = any(v is None or str(v).strip().upper() == "NULL" for v in vals.values())
    if is_null:
        note = row[36].value if len(row) > 36 else None
        loc = row[35].value if len(row) > 35 else None
        mentions_part3 = bool(loc and "제3부" in str(loc))
        null_rows.append((row[2].value, row[11].value, vals, mentions_part3, note))

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"전체 데이터 행: {total_rows}\n")
    f.write(f"수익률 3필드 중 하나라도 NULL인 행: {len(null_rows)}\n")
    mentions_part3_count = sum(1 for r in null_rows if r[3])
    f.write(f"  그 중 근거위치에 '제3부' 언급된 행: {mentions_part3_count}\n\n")
    for code, cls, vals, mentions, note in null_rows:
        f.write(f"{code} ({cls}) 제3부언급={mentions}: {vals}\n  메모: {note}\n\n")

print("done")
