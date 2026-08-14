import io
import os
import re
import sys
import glob

sys.path.insert(0, r"C:\Users\kevin\pension-agent\src")
from parsing.prospectus_tables import extract_class_returns

import openpyxl

TRACKER = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (1).xlsm"
PROSPECTUS_BASE = r"C:\Users\kevin\pension-agent\data\raw\prospectus"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_batch_extract_returns.txt"

# 폴더명(한글) 유니코드 정규화 이슈 회피: os.listdir로 실제 디렉토리명을 가져와 조합
subdir = [n for n in os.listdir(PROSPECTUS_BASE) if os.path.isdir(os.path.join(PROSPECTUS_BASE, n))][0]
PROSPECTUS_DIR = os.path.join(PROSPECTUS_BASE, subdir)

wb = openpyxl.load_workbook(TRACKER, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

tracker_rows = {}
for r in range(6, ws.max_row + 1):
    code = ws.cell(row=r, column=3).value
    if not code:
        continue
    cls = ws.cell(row=r, column=12).value
    r1 = ws.cell(row=r, column=17).value
    r3 = ws.cell(row=r, column=18).value
    rsince = ws.cell(row=r, column=19).value
    tracker_rows.setdefault(code, []).append((r, cls, r1, r3, rsince))

pdf_files = glob.glob(os.path.join(PROSPECTUS_DIR, "**", "*.pdf"), recursive=True)

results_log = []
error_log = []
stats = {"funds_ok": 0, "funds_error": 0, "classes_extracted": 0,
         "tracker_null_now_fillable": 0, "tracker_null_still_missing": 0}

for pdf_path in pdf_files:
    fname = os.path.basename(pdf_path)
    m = re.search(r"(KR[A-Z0-9]+)", fname)
    if not m:
        continue
    code = m.group(1)
    try:
        classes = extract_class_returns(pdf_path)
        stats["funds_ok"] += 1
        stats["classes_extracted"] += len(classes)
        by_code = {c.class_code: c for c in classes}

        for (r, cls, r1, r3, rsince) in tracker_rows.get(code, []):
            was_null = any(v is None or str(v).strip().upper() == "NULL" for v in (r1, r3, rsince))
            if not was_null:
                continue
            extracted = by_code.get(cls)
            if extracted and (extracted.return_1y is not None or extracted.return_3y is not None
                               or extracted.return_since_inception is not None):
                stats["tracker_null_now_fillable"] += 1
                results_log.append(
                    f"{code} ({cls}) row{r}: NULL -> 1y={extracted.return_1y} 3y={extracted.return_3y} "
                    f"since={extracted.return_since_inception}"
                )
            else:
                stats["tracker_null_still_missing"] += 1
    except Exception as e:
        stats["funds_error"] += 1
        error_log.append(f"{code} ({fname}): {type(e).__name__}: {e}")

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"pdf_files found: {len(pdf_files)}\n")
    f.write(f"=== 통계 ===\n{stats}\n\n")
    f.write(f"=== NULL -> 채울 수 있는 값 발견 ({stats['tracker_null_now_fillable']}건) ===\n")
    for line in results_log:
        f.write(line + "\n")
    f.write(f"\n=== 추출 실패한 펀드 ({len(error_log)}건) ===\n")
    for line in error_log:
        f.write(line + "\n")

print("done")
