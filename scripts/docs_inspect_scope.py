import io
import openpyxl
from collections import Counter

PATH = r"C:\Users\kevin\Downloads\docs 수정.xlsm"
OUT = r"C:\Users\kevin\pension-agent\prospectus_check\_docs_scope.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["파싱 결과"]

header = [c.value for c in ws[5]]

status_counter = Counter()
by_doc_status = {}
total = 0
for row in ws.iter_rows(min_row=6, values_only=True):
    fnum = row[0]
    if fnum is None:
        continue
    try:
        num = int(float(fnum))
    except (TypeError, ValueError):
        continue
    total += 1
    status = str(row[12] or "(empty)")
    status_counter[status] += 1
    by_doc_status.setdefault(num, Counter())[status] += 1

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write("헤더:\n")
    for i, h in enumerate(header):
        f.write(f"  {i}: {h}\n")
    f.write(f"\n총 데이터 행: {total}\n\n")
    f.write("--- 전체 상태별 집계 ---\n")
    for k, v in status_counter.most_common():
        f.write(f"{k}: {v}\n")
    f.write("\n--- 문서별 상태 집계 (검수전/미완료가 있는 문서만) ---\n")
    for docnum in sorted(by_doc_status):
        counts = by_doc_status[docnum]
        needs_review = {k: v for k, v in counts.items() if "검수완료" not in k}
        if needs_review:
            f.write(f"doc{docnum}: {dict(counts)}\n")
