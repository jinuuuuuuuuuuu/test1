import io
import openpyxl
from collections import defaultdict

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_master_consistency.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

# 마스터 후보 필드(0-idx)
MASTER_COLS = {
    1: "명칭", 3: "운용사", 4: "작성기준일", 5: "효력발생일", 6: "위험등급", 7: "모집증권종류",
    8: "투자목적", 9: "투자전략", 10: "상품분류",
    22: "비교지수", 23: "운용역성명", 24: "운용역생년", 25: "운용역직위", 26: "운용역경력",
    27: "동종수익률", 28: "매입기준", 29: "환매기준", 30: "환매대금지급", 31: "환매수수료",
    32: "과세특징", 33: "전환가능여부",
}

by_code = defaultdict(list)
for r in range(6, ws.max_row + 1):
    code = ws.cell(row=r, column=3).value
    if not code:
        continue
    row_vals = {idx: ws.cell(row=r, column=idx + 1).value for idx in MASTER_COLS}
    by_code[code].append((r, row_vals))

mismatches = []
for code, rows in by_code.items():
    if len(rows) < 2:
        continue
    base = rows[0][1]
    for r, vals in rows[1:]:
        for idx, label in MASTER_COLS.items():
            if vals[idx] != base[idx]:
                mismatches.append((code, r, label, base[idx], vals[idx]))

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"총 상품코드 수: {len(by_code)}\n")
    f.write(f"불일치 발견: {len(mismatches)}건\n\n")
    for code, r, label, base_val, val in mismatches[:50]:
        f.write(f"{code} row{r} [{label}] 불일치:\n  base={str(base_val)[:150]}\n  this={str(val)[:150]}\n\n")
