import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (1).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_prospectus1_header_kr.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["투자설명서_파싱_검수본"]

header = [c.value for c in ws[1]]

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write("=== 헤더 (컬럼 인덱스: 이름) ===\n")
    for i, h in enumerate(header):
        f.write(f"{i}: {h}\n")

    f.write("\n=== KR515302022M 관련 행 ===\n")
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        rowtext = " ".join(str(v) for v in vals if v)
        if "KR515302022M" in rowtext or "KR5153" in str(vals[1] if len(vals) > 1 else ""):
            f.write(f"\n--- row {row[0].row} ---\n")
            for i, v in enumerate(vals):
                if v is not None and str(v).strip() != "":
                    f.write(f"  [{i}:{header[i]}] {v}\n")
