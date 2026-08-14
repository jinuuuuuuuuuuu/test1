import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]투자설명서 (1).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_prospectus_col_meaning.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)

with io.open(OUT, "w", encoding="utf-8") as f:
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        f.write(f"===== {sheetname}: 첫 3행 =====\n")
        for r in range(1, 4):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 39)]
            f.write(f"row{r}: {vals}\n\n")
