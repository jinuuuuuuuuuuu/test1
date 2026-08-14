import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\docs 수정.xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_verify_batch10.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["파싱 결과"]


def find_row(chunk_id):
    for row in ws.iter_rows(min_row=6):
        if row[7].value == chunk_id:
            return row
    return None


targets = ["doc20_chunk02", "doc36_chunk04", "doc38_chunk02", "doc40_chunk01",
           "doc41_chunk03", "doc41_chunk04", "doc57_chunk00", "doc57_chunk01"]

with io.open(OUT, "w", encoding="utf-8") as f:
    for cid in targets:
        row = find_row(cid)
        if row is None:
            f.write(f"{cid}: NOT FOUND\n\n")
            continue
        f.write(f"=== {cid} (status={row[12].value}) ===\n")
        f.write(f"chunk_text:\n{row[8].value}\n\n")
        if row[10].value:
            f.write(f"표구조화텍스트:\n{row[10].value}\n\n")
        f.write("---\n\n")

    # overall status counts
    from collections import Counter
    c = Counter()
    total = 0
    for row in ws.iter_rows(min_row=6):
        if row[0].value is None:
            continue
        total += 1
        c[str(row[12].value)] += 1
    f.write(f"\n총 행 수: {total}\n")
    for k, v in c.most_common():
        f.write(f"{k}: {v}\n")
