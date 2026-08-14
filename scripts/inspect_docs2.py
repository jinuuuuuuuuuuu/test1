import io
import openpyxl

PATH = r"C:\Users\kevin\Downloads\[파싱]renamed폴더 (2).xlsm"
OUT = r"C:\Users\kevin\pension-agent\docs_check\_docs2_inspect.txt"

wb = openpyxl.load_workbook(PATH, data_only=True)

with io.open(OUT, "w", encoding="utf-8") as f:
    f.write(f"시트 목록: {wb.sheetnames}\n\n")
    ws = wb["파싱 결과_검수본"]
    f.write(f"dims={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}\n\n")
    for r in range(1, 7):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 14)]
        f.write(f"row{r}: {vals}\n\n")
